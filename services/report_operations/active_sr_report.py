"""Configuration-driven Active Service Request report builder.

Produces two per-manager Excel workbooks (WEST = Mabi phases 3/4/4c only,
EAST = Robert phases 5/7/8 only) from enriched work-order row dicts
(work_order_validator_service.validate).

Sheet structure mirrors the original Python-in-Excel workbooks exactly:
  ServiceRequest  – raw dump (all columns); WEST → phases 3/4/4c, EAST → phases 5/7/8
  Full_unassign   – phase/unassigned overview, side-by-side tables
  By tech         – manager + team In-progress, side-by-side
  unassign        – unassigned by sub-phase, side-by-side
  Phase N         – per-phase tech breakdown, side-by-side

All filtered tables share the same five output columns, sorted Days Open desc.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.work_order_excel import (
    _HEADER_LABELS,
    _safe_val,
    _write_flat_sheet,
)

# ---------------------------------------------------------------------------
# Layout constants
# ---------------------------------------------------------------------------

# 1-based column indices for the 7 possible side-by-side table anchors:
# B=2, H=8, N=14, T=20, Z=26, AF=32, AL=38  (6-col stride: 5 data + 1 gap)
_ANCHOR_COLS: list[int] = [2, 8, 14, 20, 26, 32, 38]

# The five output columns for every filtered table (source_key, header_label)
_TABLE_COLS: list[tuple[str, str]] = [
    ("Location",    "Unit"),
    ("Assigned to", "Assigned To"),
    ("Days open",   "Days Open"),
    ("Issue",       "Issue"),
    ("Status",      "Status"),
]

# Fixed column widths matching the five output columns
_COL_WIDTHS: list[int] = [12, 20, 10, 35, 14]

# Palette — section title rows + column headers (both reports, all sheets)
_C_HEADER_BLACK = "000000"
_C_WHITE = "FFFFFF"

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class FilterParams:
    """Declarative filter specification for one table section.

    phases          – frozenset of phase strings to keep (exact match after strip;
                      compared uppercase). None = no phase filter (raw_dump only).
    assigned        – exact assigned-to name match (case-insensitive strip).
                      None = no assigned filter.
    status          – exact status match (case-insensitive strip).
                      None = no status filter.
    unassigned_mode – "broad"  → keep rows where Assigned to ∈ {"Unassigned", ""}
                      "strict" → keep rows where Assigned to == "Unassigned"
                      None     → no unassigned filter
    assigned and unassigned_mode are mutually exclusive.
    """

    phases: frozenset[str] | None = None
    assigned: str | None = None
    status: str | None = None
    unassigned_mode: str | None = None

    def __post_init__(self) -> None:
        if self.assigned is not None and self.unassigned_mode is not None:
            raise ValueError(
                "FilterParams: 'assigned' and 'unassigned_mode' are mutually exclusive."
            )
        if self.unassigned_mode not in (None, "broad", "strict"):
            raise ValueError(
                f"FilterParams: unassigned_mode must be 'broad', 'strict', or None; "
                f"got {self.unassigned_mode!r}"
            )


@dataclass(frozen=True)
class SectionDef:
    """One horizontal table block on a sheet."""

    title: str          # text for the merged title cell(s) above the table
    filter: FilterParams


@dataclass(frozen=True)
class SheetDef:
    """One worksheet definition.

    tab_name  – exact Excel tab name (trailing spaces are preserved).
    sections  – ordered list; engine places them left-to-right at anchor cols.
    title_row – first row of the 2-row merged title block (ignored for raw_dump).
    data_row  – row for column headers (data starts at data_row + 1).
    kind      – "horizontal" (default) or "raw_dump" (ServiceRequest sheet).
    phases    – optional phase filter applied to raw_dump sheets before writing.
                None = no filtering (all rows passed through).
    """

    tab_name: str
    sections: list[SectionDef]
    title_row: int
    data_row: int
    kind: str = "horizontal"
    phases: frozenset[str] | None = None
    classification: str | None = None


@dataclass(frozen=True)
class ReportConfig:
    """Top-level report definition."""

    report_name: str
    sheets: list[SheetDef]


# Portfolio boundaries — EAST/WEST horizontal sections and ServiceRequest raw_dump use these sets.
_PHASES_WEST = frozenset({"3", "4", "4c"})
_PHASES_EAST = frozenset({"5", "7", "8"})


# ---------------------------------------------------------------------------
# Filter engine
# ---------------------------------------------------------------------------


def _safe_days(row: dict) -> int:
    try:
        return int(row.get("Days open") or 0)
    except (TypeError, ValueError):
        return 0


def _filter_rows(rows: list[dict], fp: FilterParams) -> list[dict]:
    """Apply FilterParams to rows and return sorted (Days open desc) subset."""
    result = rows

    # Phase filter — strict: row ph must exactly match one allowed token (after
    # strip + uppercase). Rows with missing/blank ph never pass.
    if fp.phases is not None:
        allowed = frozenset(
            str(p).strip().upper() for p in fp.phases if str(p).strip()
        )
        if not allowed:
            result = []
        else:
            result = [
                r
                for r in result
                if (t := str(r.get("ph", "")).strip().upper()) != "" and t in allowed
            ]

    # Assigned / unassigned filter
    if fp.unassigned_mode == "broad":
        result = [
            r for r in result
            if str(r.get("Assigned to") or "").strip() in ("Unassigned", "")
        ]
    elif fp.unassigned_mode == "strict":
        result = [
            r for r in result
            if str(r.get("Assigned to") or "").strip() == "Unassigned"
        ]
    elif fp.assigned is not None:
        target = fp.assigned.lower()
        result = [
            r for r in result
            if str(r.get("Assigned to") or "").strip().lower() == target
        ]

    # Status filter
    if fp.status is not None:
        target_s = fp.status.lower()
        result = [
            r for r in result
            if str(r.get("Status") or "").strip().lower() == target_s
        ]

    # Sort Days open descending
    result = sorted(result, key=_safe_days, reverse=True)
    return result


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _bold_white() -> Font:
    return Font(bold=True, color=_C_WHITE)


# ---------------------------------------------------------------------------
# Horizontal layout engine
# ---------------------------------------------------------------------------


def _write_title_block(
    ws,
    col_start: int,
    title_row: int,
    data_row: int,
    title: str,
) -> None:
    """Write merged title cell(s) spanning title_row to data_row-1."""
    title_end_row = data_row - 1
    col_end = col_start + len(_TABLE_COLS) - 1

    start_cell = ws.cell(row=title_row, column=col_start)
    end_letter = get_column_letter(col_end)

    ws.merge_cells(
        start_row=title_row,
        start_column=col_start,
        end_row=title_end_row,
        end_column=col_end,
    )
    start_cell.value = title
    start_cell.font = _bold_white()
    start_cell.fill = _fill(_C_HEADER_BLACK)
    start_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_table_headers(ws, col_start: int, header_row: int) -> None:
    """Write the five column header cells."""
    for offset, (_, label) in enumerate(_TABLE_COLS):
        cell = ws.cell(row=header_row, column=col_start + offset, value=label)
        cell.font = _bold_white()
        cell.fill = _fill(_C_HEADER_BLACK)
        cell.alignment = _center()


def _write_table_data(
    ws,
    col_start: int,
    data_start_row: int,
    filtered_rows: list[dict],
) -> None:
    """Write filtered data rows starting at data_start_row."""
    for row_offset, row in enumerate(filtered_rows):
        excel_row = data_start_row + row_offset
        for col_offset, (key, _) in enumerate(_TABLE_COLS):
            val = _safe_val(row, key)
            cell = ws.cell(row=excel_row, column=col_start + col_offset, value=val)
            cell.alignment = _center()


def _set_column_widths(ws, col_start: int) -> None:
    """Apply fixed widths to the five columns of one table block."""
    for offset, width in enumerate(_COL_WIDTHS):
        letter = get_column_letter(col_start + offset)
        existing = ws.column_dimensions[letter].width
        ws.column_dimensions[letter].width = max(existing or 0, width)


def _write_horizontal_sheet(
    ws,
    sheet_def: SheetDef,
    rows: list[dict],
) -> None:
    """Render all sections side-by-side on a single worksheet."""
    for idx, section in enumerate(sheet_def.sections):
        col_start = _ANCHOR_COLS[idx]
        title_row = sheet_def.title_row
        data_row = sheet_def.data_row

        _write_title_block(ws, col_start, title_row, data_row, section.title)
        _write_table_headers(ws, col_start, data_row)

        filtered = _filter_rows(rows, section.filter)
        _write_table_data(ws, col_start, data_row + 1, filtered)
        _set_column_widths(ws, col_start)


# ---------------------------------------------------------------------------
# Report renderer
# ---------------------------------------------------------------------------


def _render_report(config: ReportConfig, rows: list[dict]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for sheet_def in config.sheets:
        ws = wb.create_sheet(sheet_def.tab_name)
        if sheet_def.kind == "raw_dump":
            rows_to_write = (
                _filter_rows(rows, FilterParams(phases=sheet_def.phases))
                if sheet_def.phases is not None
                else rows
            )
            if sheet_def.classification is not None:
                clsf = sheet_def.classification.lower()
                rows_to_write = [
                    r for r in rows_to_write
                    if str(r.get("wo_classification") or "").lower() == clsf
                ]
            _write_flat_sheet(ws, rows_to_write, header_bg=_C_HEADER_BLACK)
            num_cols = len(_HEADER_LABELS)
            ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"
        else:
            _write_horizontal_sheet(ws, sheet_def, rows)

    return wb


# ---------------------------------------------------------------------------
# WEST_CONFIG — Mabi (phases 3 / 4 / 4c)
# ---------------------------------------------------------------------------

WEST_CONFIG = ReportConfig(
    report_name="WEST",
    sheets=[

        # Sheet 1: raw data source — filtered to WEST phases only
        SheetDef(
            tab_name="ServiceRequest",
            sections=[],
            title_row=1,
            data_row=1,
            kind="raw_dump",
            phases=_PHASES_WEST,
        ),

        # Sheet 2: cross-phase overview (5 tables, title rows 3–4, data row 5)
        SheetDef(
            tab_name="Full_unassign",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Full Service request  Report -Mabi Side",
                    filter=FilterParams(phases=_PHASES_WEST),
                ),
                SectionDef(
                    title="Full Service request  Report Unassigned Mabi",
                    filter=FilterParams(unassigned_mode="strict"),
                ),
                SectionDef(
                    title="Unassigned Ph 3",
                    filter=FilterParams(phases=frozenset({"3"}), unassigned_mode="broad"),
                ),
                SectionDef(
                    title="Unassigned Ph 4",
                    filter=FilterParams(phases=frozenset({"4"}), unassigned_mode="broad"),
                ),
                SectionDef(
                    title="Unassigned Ph 4c",
                    filter=FilterParams(phases=frozenset({"4c"}), unassigned_mode="broad"),
                ),
            ],
        ),

        # Sheet 3: manager + team In-progress (6 tables, title rows 2–3, data row 4)
        # Source scripts #6–#11: assignee/status only — no phase filter.
        SheetDef(
            tab_name="By tech ",          # trailing space is intentional
            title_row=2,
            data_row=4,
            sections=[
                SectionDef(
                    title="Mabi",
                    filter=FilterParams(assigned="Mabi"),
                ),
                SectionDef(
                    title="Dennis Arevalo",
                    filter=FilterParams(
                        assigned="Dennis Arevalo",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Rayniel Rincon",
                    filter=FilterParams(
                        assigned="Rayniel Rincon",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Alexander Gonzalez",
                    filter=FilterParams(
                        assigned="Alexander Gonzalez",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Diego Zapata",
                    filter=FilterParams(
                        assigned="Diego Zapata",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Victor Castaneda",
                    filter=FilterParams(
                        assigned="Victor Castaneda",
                        status="In progress",
                    ),
                ),
            ],
        ),

        # Sheet 4: unassigned index by sub-phase (3 tables, title rows 2–3, data row 4)
        SheetDef(
            tab_name="unassign",
            title_row=2,
            data_row=4,
            sections=[
                SectionDef(
                    title="Phase 3",
                    filter=FilterParams(phases=frozenset({"3"}), unassigned_mode="broad"),
                ),
                SectionDef(
                    title="Phase 4",
                    filter=FilterParams(phases=frozenset({"4"}), unassigned_mode="broad"),
                ),
                SectionDef(
                    title="Phase 4c",
                    filter=FilterParams(phases=frozenset({"4c"}), unassigned_mode="broad"),
                ),
            ],
        ),

        # Sheet 5: Phase 3 deep-dive (7 tables, title rows 3–4, data row 5)
        # Sections 1–2: source scripts #9/#14 — assignee+status only, no phase filter.
        # Sections 4–7: source scripts #16–#19 — include phase 3 filter.
        SheetDef(
            tab_name="Phase 3",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Alexander Gonzalez",
                    filter=FilterParams(
                        assigned="Alexander Gonzalez",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Alexander Gonzalez",
                    filter=FilterParams(
                        assigned="Alexander Gonzalez",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Alexander Gonzalez/Victor Castaneda/Dennis Arevalo",
                    filter=FilterParams(phases=frozenset({"3"}), unassigned_mode="strict"),
                ),
                SectionDef(
                    title="Victor Castaneda",
                    filter=FilterParams(
                        phases=frozenset({"3"}),
                        assigned="Victor Castaneda",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Victor Castaneda",
                    filter=FilterParams(
                        phases=frozenset({"3"}),
                        assigned="Victor Castaneda",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Dennis Arevalo",
                    filter=FilterParams(
                        phases=frozenset({"3"}),
                        assigned="Dennis Arevalo",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Dennis Arevalo",
                    filter=FilterParams(
                        phases=frozenset({"3"}),
                        assigned="Dennis Arevalo",
                        status="On hold",
                    ),
                ),
            ],
        ),

        # Sheet 6: Phase 4 deep-dive (7 tables)
        # Sections 1–2: source scripts #8/#20 — assignee+status only, no phase filter.
        # Sections 4–7: source scripts #22–#25 — include phase 4 filter.
        SheetDef(
            tab_name="Phase 4",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Rayniel Rincon",
                    filter=FilterParams(
                        assigned="Rayniel Rincon",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Rayniel Rincon",
                    filter=FilterParams(
                        assigned="Rayniel Rincon",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Rayniel Rincon/Diego Zapata/Dennis Arevalo",
                    filter=FilterParams(phases=frozenset({"4"}), unassigned_mode="strict"),
                ),
                SectionDef(
                    title="Diego Zapata",
                    filter=FilterParams(
                        phases=frozenset({"4"}),
                        assigned="Diego Zapata",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Diego Zapata",
                    filter=FilterParams(
                        phases=frozenset({"4"}),
                        assigned="Diego Zapata",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Dennis Arevalo",
                    filter=FilterParams(
                        phases=frozenset({"4"}),
                        assigned="Dennis Arevalo",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Dennis Arevalo",
                    filter=FilterParams(
                        phases=frozenset({"4"}),
                        assigned="Dennis Arevalo",
                        status="On hold",
                    ),
                ),
            ],
        ),

        # Sheet 7: Phase 4c deep-dive (3 tables — only Victor covers this phase)
        SheetDef(
            tab_name="Phase 4c",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Victor Castaneda",
                    filter=FilterParams(
                        phases=frozenset({"4c"}),
                        assigned="Victor Castaneda",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Victor Castaneda",
                    filter=FilterParams(
                        phases=frozenset({"4c"}),
                        assigned="Victor Castaneda",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Victor Castaneda",
                    filter=FilterParams(phases=frozenset({"4c"}), unassigned_mode="strict"),
                ),
            ],
        ),

        # Sheet 8: Make Ready — all Make Ready WOs in WEST phases (flat, all columns)
        SheetDef(
            tab_name="Make Ready",
            sections=[],
            title_row=1,
            data_row=1,
            kind="raw_dump",
            phases=_PHASES_WEST,
            classification="Make Ready",
        ),
    ],
)


# ---------------------------------------------------------------------------
# EAST_CONFIG — Robert (phases 5 / 7 / 8 only)
# ---------------------------------------------------------------------------

EAST_CONFIG = ReportConfig(
    report_name="EAST",
    sheets=[

        # Sheet 1: raw data source — Robert portfolio phases only (matches WEST scoping)
        SheetDef(
            tab_name="ServiceRequest",
            sections=[],
            title_row=1,
            data_row=1,
            kind="raw_dump",
            phases=_PHASES_EAST,
        ),

        # Sheet 2: portfolio overview (6 tables, title rows 3–4, data row 5)
        # Cols 1–2: EAST-phase full dump + unassigned; 3–4: EAST; 5–6: Mabi cross-view 3/4/4c.
        SheetDef(
            tab_name="Full_unassign",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Full Service request  Report ",
                    filter=FilterParams(phases=_PHASES_EAST),
                ),
                SectionDef(
                    title="Full Service request  Report Unassigned",
                    filter=FilterParams(phases=_PHASES_EAST, unassigned_mode="strict"),
                ),
                SectionDef(
                    title="Full Service request  Report -Roberto Side",
                    filter=FilterParams(phases=_PHASES_EAST),
                ),
                SectionDef(
                    title="Full Service request  Report Unassigned Robertos ",
                    filter=FilterParams(phases=_PHASES_EAST, unassigned_mode="broad"),
                ),
                SectionDef(
                    title="Full Service request  Report -Mabi Side",
                    filter=FilterParams(phases=frozenset({"3", "4", "4c"})),
                ),
                SectionDef(
                    title="Full Service request  Report Unassigned Mabi",
                    filter=FilterParams(phases=frozenset({"3", "4", "4c"}), unassigned_mode="broad"),
                ),
            ],
        ),

        # Sheet 3: manager + team In-progress (5 tables — Robert has 4 techs vs Mabi's 5)
        SheetDef(
            tab_name="By tech ",          # trailing space intentional
            title_row=2,
            data_row=4,
            sections=[
                SectionDef(
                    title="Roberto Palacios",
                    filter=FilterParams(phases=_PHASES_EAST, assigned="Roberto Palacios"),
                ),
                SectionDef(
                    title="Latrell Dawson",
                    filter=FilterParams(
                        phases=_PHASES_EAST,
                        assigned="Latrell Dawson",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Yomar Gonzalez",
                    filter=FilterParams(
                        phases=_PHASES_EAST,
                        assigned="Yomar Gonzalez",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Barron Russell",
                    filter=FilterParams(
                        phases=_PHASES_EAST,
                        assigned="Barron Russell",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Antonio Sherfield",
                    filter=FilterParams(
                        phases=_PHASES_EAST,
                        assigned="Antonio Sherfield",
                        status="In progress",
                    ),
                ),
            ],
        ),

        # Sheet 4: unassigned index by sub-phase (3 tables — 5 / 7 / 8 only)
        SheetDef(
            tab_name="unassign",
            title_row=2,
            data_row=4,
            sections=[
                SectionDef(
                    title="Phase 5",
                    filter=FilterParams(phases=frozenset({"5"}), unassigned_mode="broad"),
                ),
                SectionDef(
                    title="Phase 7",
                    filter=FilterParams(phases=frozenset({"7"}), unassigned_mode="broad"),
                ),
                SectionDef(
                    title="Phase 8",
                    filter=FilterParams(phases=frozenset({"8"}), unassigned_mode="broad"),
                ),
            ],
        ),

        # Sheet 5: Phase 5 (5 tables — Latrell + Yomar)
        SheetDef(
            tab_name="Phase 5",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Latrell Dawson",
                    filter=FilterParams(
                        phases=frozenset({"5"}),
                        assigned="Latrell Dawson",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Latrell Dawson",
                    filter=FilterParams(
                        phases=frozenset({"5"}),
                        assigned="Latrell Dawson",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Latrell Dawson/Yomar Gonzalez",
                    filter=FilterParams(phases=frozenset({"5"}), unassigned_mode="strict"),
                ),
                SectionDef(
                    title="Yomar Gonzalez",
                    filter=FilterParams(
                        phases=frozenset({"5"}),
                        assigned="Yomar Gonzalez",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Yomar Gonzalez",
                    filter=FilterParams(
                        phases=frozenset({"5"}),
                        assigned="Yomar Gonzalez",
                        status="On hold",
                    ),
                ),
            ],
        ),

        # Sheet 6: Phase 7 (3 tables — Barron only)
        SheetDef(
            tab_name="Phase 7",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Barron Russell",
                    filter=FilterParams(
                        phases=frozenset({"7"}),
                        assigned="Barron Russell",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Barron Russell",
                    filter=FilterParams(
                        phases=frozenset({"7"}),
                        assigned="Barron Russell",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Barron Russell",
                    filter=FilterParams(phases=frozenset({"7"}), unassigned_mode="strict"),
                ),
            ],
        ),

        # Sheet 7: Phase 8 (3 tables — Antonio only)
        SheetDef(
            tab_name="Phase 8",
            title_row=3,
            data_row=5,
            sections=[
                SectionDef(
                    title="Antonio Sherfield",
                    filter=FilterParams(
                        phases=frozenset({"8"}),
                        assigned="Antonio Sherfield",
                        status="In progress",
                    ),
                ),
                SectionDef(
                    title="Antonio Sherfield",
                    filter=FilterParams(
                        phases=frozenset({"8"}),
                        assigned="Antonio Sherfield",
                        status="On hold",
                    ),
                ),
                SectionDef(
                    title="Antonio Sherfield",
                    filter=FilterParams(phases=frozenset({"8"}), unassigned_mode="strict"),
                ),
            ],
        ),

        # Sheet 8: Make Ready — all Make Ready WOs in EAST phases (flat, all columns)
        SheetDef(
            tab_name="Make Ready",
            sections=[],
            title_row=1,
            data_row=1,
            kind="raw_dump",
            phases=_PHASES_EAST,
            classification="Make Ready",
        ),
    ],
)


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------

_CONFIGS: dict[str, ReportConfig] = {
    "WEST": WEST_CONFIG,
    "EAST": EAST_CONFIG,
}


def build_active_sr_report_from_rows(
    rows: list[dict],
    report: str = "WEST",
) -> bytes:
    """Build a manager report workbook from already-validated rows.

    Use this when rows have already been produced by
    work_order_validator_service.validate() — avoids double-parsing.

    Args:
        rows:   enriched row dicts from validate()
        report: "WEST" (Mabi) or "EAST" (Robert)

    Returns:
        Raw bytes of the .xlsx workbook.
    """
    key = report.upper()
    if key not in _CONFIGS:
        raise ValueError(f"Unknown report name {report!r}. Expected 'WEST' or 'EAST'.")
    config = _CONFIGS[key]
    wb = _render_report(config, rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_active_sr_report(
    property_id: int,
    sr_file_content: bytes,
    report: str = "WEST",
) -> bytes:
    """Parse, classify, and render a manager report workbook.

    Calls work_order_validator_service.validate() internally.
    Prefer build_active_sr_report_from_rows() when rows are already computed
    to avoid double DB hits.

    Args:
        property_id:      tenant-scoped property identifier
        sr_file_content:  raw bytes of the uploaded Active SR Excel file
        report:           "WEST" (Mabi) or "EAST" (Robert)

    Returns:
        Raw bytes of the .xlsx workbook.
    """
    from services.work_order_validator_service import validate  # avoid circular at module load

    rows = validate(property_id, sr_file_content)
    return build_active_sr_report_from_rows(rows, report)
