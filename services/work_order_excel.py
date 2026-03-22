"""Excel builder for the Work Order Validator report.

Produces a multi-sheet workbook:
  - "{Manager} – All"        : all WOs for that manager's phases
  - "{Manager} – Unassigned" : WOs with no assignee
  - "{Manager} – {Assignee}" : one sheet per technician / assignee
  - "Other"                  : common areas + unrecognized phases

Within each per-assignee sheet:
  IN PROGRESS section (blue header row) + ON HOLD section (orange header row)

WO Classification cell coloring:
  Make Ready              → green (#C6EFCE)
  Service Technician      → no fill
  Service Tech – Common … → light gray (#D9D9D9)
"""

from __future__ import annotations

import io
from datetime import date as _date
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Manager phase sets (uppercased)
# ---------------------------------------------------------------------------
MABI_PHASES: frozenset[str] = frozenset({"3", "4", "4C", "4S"})
ROBERT_PHASES: frozenset[str] = frozenset({"5", "7", "8"})

# ---------------------------------------------------------------------------
# Output column definitions
# ---------------------------------------------------------------------------
_COLUMNS = [
    ("ph",               "PH"),
    ("bld",              "BLD"),
    ("Days open",        "Days Open"),
    ("Number",           "Number"),
    ("Location",         "Location"),
    ("Created date",     "Created Date"),
    ("Due date",         "Due Date"),
    ("Service Category", "Service Category"),
    ("Issue",            "Issue"),
    ("Assigned to",      "Assigned To"),
    ("Priority",         "Priority"),
    ("Status",           "Status"),
    ("wo_classification","WO Classification"),
]

_FIELD_KEYS = [c[0] for c in _COLUMNS]
_HEADER_LABELS = [c[1] for c in _COLUMNS]

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
_C_HEADER_BG    = "1F497D"   # dark navy — header row background
_C_SECTION_IP   = "BDD7EE"   # light blue — IN PROGRESS section header
_C_SECTION_OH   = "FCE4D6"   # light orange — ON HOLD section header
_C_MR           = "C6EFCE"   # green — Make Ready
_C_UNRESOLVED   = "D9D9D9"   # gray — Common Area
_DATE_FMT       = "MM/DD/YY"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _bold(color: str = "000000") -> Font:
    return Font(bold=True, color=color)


def _classification_fill(classification: str) -> PatternFill | None:
    c = str(classification)
    if c == "Make Ready":
        return _fill(_C_MR)
    if "Common Area" in c:
        return _fill(_C_UNRESOLVED)
    return None  # Service Technician — no fill


def _safe_val(row: dict, key: str) -> Any:
    val = row.get(key)
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    if isinstance(val, pd.Timestamp):
        return val.date()
    return val


def _group_for_phase(phase: str) -> str:
    """Return 'Robert', 'Mabi', or 'Other' for a given phase string."""
    p = str(phase).upper()
    if p in ROBERT_PHASES:
        return "Robert"
    if p in MABI_PHASES:
        return "Mabi"
    return "Other"


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_header_row(ws, row_num: int, *, header_bg: str | None = None) -> None:
    """Write the column header row with styling."""
    bg = header_bg if header_bg is not None else _C_HEADER_BG
    for col_idx, label in enumerate(_HEADER_LABELS, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill(bg)
        cell.alignment = _center()


def _write_data_row(ws, row_num: int, row: dict) -> None:
    """Write one WO data row with per-cell styling."""
    classification = str(row.get("wo_classification") or "")
    clf_fill = _classification_fill(classification)
    clf_col_idx = _FIELD_KEYS.index("wo_classification") + 1

    for col_idx, key in enumerate(_FIELD_KEYS, start=1):
        val = _safe_val(row, key)
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.alignment = _center()
        if isinstance(val, _date):
            cell.number_format = _DATE_FMT
        if col_idx == clf_col_idx and clf_fill:
            cell.fill = clf_fill


def _write_section_header(ws, row_num: int, label: str, bg_color: str) -> None:
    """Write a full-width section label row (merged, bold)."""
    num_cols = len(_COLUMNS)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A{row_num}:{last_col}{row_num}")
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = Font(bold=True, color="000000", size=11)
    cell.fill = _fill(bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _autofit_columns(ws, rows: list[dict]) -> None:
    """Set column widths based on header and data lengths."""
    for col_idx, (key, label) in enumerate(_COLUMNS, start=1):
        max_len = len(label)
        for row in rows:
            val = _safe_val(row, key)
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def _write_flat_sheet(
    ws,
    rows: list[dict],
    *,
    header_bg: str | None = None,
) -> None:
    """Write a header row followed by all data rows (no section splits).

    header_bg: optional Excel ARGB (no #). When None, uses default navy _C_HEADER_BG.
    """
    _write_header_row(ws, row_num=1, header_bg=header_bg)
    for i, row in enumerate(rows, start=2):
        _write_data_row(ws, i, row)
    ws.freeze_panes = "A2"
    _autofit_columns(ws, rows)


def _write_technician_sheet(ws, in_progress: list[dict], on_hold: list[dict]) -> None:
    """Write IN PROGRESS section + blank row + ON HOLD section."""
    current_row = 1

    # IN PROGRESS section
    _write_section_header(ws, current_row, "IN PROGRESS", _C_SECTION_IP)
    current_row += 1
    _write_header_row(ws, current_row)
    current_row += 1
    for row in in_progress:
        _write_data_row(ws, current_row, row)
        current_row += 1

    # Blank separator
    current_row += 1

    # ON HOLD section
    _write_section_header(ws, current_row, "ON HOLD", _C_SECTION_OH)
    current_row += 1
    _write_header_row(ws, current_row)
    current_row += 1
    for row in on_hold:
        _write_data_row(ws, current_row, row)
        current_row += 1

    _autofit_columns(ws, in_progress + on_hold)


# ---------------------------------------------------------------------------
# Safe sheet name helper
# ---------------------------------------------------------------------------

def _sheet_name(text: str, max_len: int = 31) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    invalid = r'\/:*?[]'
    clean = "".join(c for c in text if c not in invalid)
    return clean[:max_len]


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_work_order_report(rows: list[dict]) -> bytes:
    """Build the multi-sheet WO Validator Excel workbook and return bytes."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Sort all rows by Days Open descending
    def _days_sort_key(r: dict) -> int:
        val = r.get("Days open")
        try:
            return -int(val)
        except (TypeError, ValueError):
            return 0

    rows_sorted = sorted(rows, key=_days_sort_key)

    # Partition by manager group
    robert_rows = [r for r in rows_sorted if _group_for_phase(r.get("ph", "")) == "Robert"]
    mabi_rows   = [r for r in rows_sorted if _group_for_phase(r.get("ph", "")) == "Mabi"]
    other_rows  = [r for r in rows_sorted if _group_for_phase(r.get("ph", "")) == "Other"]

    def _is_unassigned(r: dict) -> bool:
        val = str(r.get("Assigned to") or "").strip()
        return val in ("", "Unassigned")

    # Build sheets for one manager group
    def _add_group_sheets(manager: str, group_rows: list[dict]) -> None:
        if not group_rows:
            return

        # All sheet
        ws_all = wb.create_sheet(_sheet_name(f"{manager} – All"))
        _write_flat_sheet(ws_all, group_rows)

        # Unassigned sheet
        unassigned = [r for r in group_rows if _is_unassigned(r)]
        ws_un = wb.create_sheet(_sheet_name(f"{manager} – Unassigned"))
        _write_flat_sheet(ws_un, unassigned)

        # Per-assignee sheets (alphabetically sorted, excluding "Unassigned"/empty)
        assignees = sorted({
            str(r.get("Assigned to") or "").strip()
            for r in group_rows
            if not _is_unassigned(r)
        })
        for assignee in assignees:
            tech_rows = [r for r in group_rows if str(r.get("Assigned to") or "").strip() == assignee]
            in_progress = [r for r in tech_rows if str(r.get("Status") or "").strip() == "In progress"]
            on_hold     = [r for r in tech_rows if str(r.get("Status") or "").strip() == "On hold"]
            ws = wb.create_sheet(_sheet_name(f"{manager} – {assignee}"))
            _write_technician_sheet(ws, in_progress, on_hold)

    _add_group_sheets("Robert", robert_rows)
    _add_group_sheets("Mabi", mabi_rows)

    # Other sheet
    if other_rows:
        ws_other = wb.create_sheet("Other")
        _write_flat_sheet(ws_other, other_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
